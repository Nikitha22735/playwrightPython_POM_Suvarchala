import csv
import json
import os

from dotenv import load_dotenv
from openpyxl import load_workbook
import pytest


def jsonHandling(filepath):
    with open(filepath) as data:
        formatedData = json.load(data)

    return formatedData


def test_json():
    result = jsonHandling("data/creds.json")



def csvHandling():
    with open("data/credentails.csv") as data:
        formatedData = csv.DictReader(data)
        dataList = []
        for i in formatedData:
            dataList.append(i)
        print(dataList)



def csvWrite():
    with open("data/credentails.csv", mode='a', newline="") as data:
        formatedData = csv.DictWriter(data, fieldnames=["username","password"])
        formatedData.writerow({'username': 'tripur123', 'password': 'welcome'})


# pip install openpyxl
def excelhandling():
        
        workbook = load_workbook("data\\sample_creds.xlsx")
        sheet = workbook["sheet1"]
        data = []
        for i in sheet.iter_rows(min_row=2, values_only=True):
            data.append(i)

        print(data)

# set useremail=test123&&set pwText=passwordvalue&&pytest -m dataHandling -s
# @pytest.mark.dataHandling
def test_CLI():
    username = os.getenv("useremail")
    pw = os.getenv("pwText")
    print(username)
    print(pw)


#pip install python-dotenv
@pytest.mark.dataHandling
def test_CLI():
    load_dotenv(os.getenv("envFile"))
    print(os.getenv("url"))
    print(os.getenv("usename"))
    print(os.getenv("pw"))



